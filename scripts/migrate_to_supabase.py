"""
Migracion one-time (Fase A) del Excel legado a Supabase.

Lee "REQUERIMIENTOS BOLSAS DE HORAS 414.xlsx" (Dashboard Principal + 7 hojas
de detalle) e inserta/actualiza projects/requirements/requirement_tasks en
Supabase. Requiere que supabase/schema.sql ya se haya corrido en el proyecto
(la fila del proyecto "positiva-web-414" ya debe existir por la seed del DDL).

No pobla activity_logs ni document_versions (no hay fuente 1:1 en el Excel).
planned_start_date/planned_end_date quedan NULL para todas las tareas: se
inspeccionaron las 4 hojas Gantt ocultas y el match por nombre de tarea no es
viable (nombres genericos, no coinciden con task_name de las hojas de
detalle) -- ver ROADMAP_SUPABASE.md SS4.4. El Gantt (/planeacion) usa
due_date como fallback.

ADVERTENCIA (desde Fase C): --reset hace DELETE FROM requirements, y eso
CASCADEA a activity_logs y document_versions. Esas tablas contienen datos
que a partir de Fase C no existen en ninguna otra parte (bitacora de horas,
documentos). NO usar --reset salvo que se acepte perder ese historial.
Nota aparte: el .xlsx fuente de este script fue borrado (ver CLAUDE.md), asi
que --reset ya no es re-ejecutable de todas formas -- no hay de donde volver
a leer los datos.

Uso:
    pip install openpyxl supabase
    $env:SUPABASE_URL = "https://<project-ref>.supabase.co"
    $env:SUPABASE_SECRET_KEY = "sb_secret_..."   # equivalente a service_role, bypassa RLS
    python migrate_to_supabase.py
    python migrate_to_supabase.py --reset   # borra requirements del proyecto antes de reinsertar

Usa el cliente supabase-py (API REST/PostgREST) en vez de una conexion
directa a Postgres -- no requiere la contrasena de base de datos, solo la
secret key del proyecto (Project Settings > API). El schema se aplica aparte
vía las migraciones versionadas en supabase/migrations/ (npm run db:push,
ver supabase/MIGRACIONES.md) -- este script asume que las tablas y la seed
del proyecto ya existen.

Idempotente: upsert por clave natural (project_id+code para requirements,
requirement_id+phase_number+task_name para tareas). --reset hace DELETE por
project_id antes de reinsertar, para partir de cero sin duplicar filas
huerfanas si el mapeo de columnas cambio entre corridas.
"""

import os
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
from supabase import create_client

EXCEL_PATH = Path(__file__).resolve().parent.parent.parent / "legado" / "REQUERIMIENTOS BOLSAS DE HORAS 414.xlsx"
PROJECT_SLUG = "positiva-web-414"
DASHBOARD_SHEET = "Dashboard Principal"

FASES_ORDEN = [
    (1, "Requerimientos"),
    (2, "Diseño"),
    (3, "Desarrollo"),
    (4, "QA"),
    (5, "Producción"),
]

ESTADO_ES_A_ENUM = {
    "en curso": "EN_CURSO",
    "pausado": "PAUSADO",
    "no iniciado": "NO_INICIADO",
    "entregado en producción": "ENTREGADO_PRODUCCION",
    "entregado en produccion": "ENTREGADO_PRODUCCION",
}

# Mismo respaldo que src/lib/excel/dashboard-sheet.ts (ESTADO_HEURISTICO) --
# los 21 requerimientos sin hoja de detalle propia, estado recuperado de un
# backup antes de que se limpiara del Excel. Se mantiene aqui como fuente
# unica de verdad para la migracion (no se importa el TS desde Python).
ESTADO_HEURISTICO = {
    "SALUD_HU0001_ModificaciónPantalla+Salud": "ENTREGADO_PRODUCCION",
    "ADMI_HU0001_AdministradorPortalPensionados": "ENTREGADO_PRODUCCION",
    "INTE_HU0001_MODULO_DE_BUSQUEDA_INTERMEDIARIOS": "ENTREGADO_PRODUCCION",
    "ACRO_HU0001_ModificacionDocumentosAcrobat": "ENTREGADO_PRODUCCION",
    "ACRO_HU0002_ModificacionDocumentosAcrobat": "ENTREGADO_PRODUCCION",
    "GEWE_HU0006_GestiónDocumentos": "ENTREGADO_PRODUCCION",
    "PARQ_HU0001_DiseñoPáginaInterna": "ENTREGADO_PRODUCCION",
    "COTI_HU0001_LevantamientoDeRequerimiento": "ENTREGADO_PRODUCCION",
    "Estandarización Documental": "ENTREGADO_PRODUCCION",
    "Desarrollo y puesta en producción - Exequias": "ENTREGADO_PRODUCCION",
    "Produccion VIGILANTES CORPORATIVOS EXEQUIAS ARL": "ENTREGADO_PRODUCCION",
    "Wompi (FR14) Validaciones Cambios Gaia": "ENTREGADO_PRODUCCION",
    "ACRO_HU0005_RediseñoFurelFurat": "ENTREGADO_PRODUCCION",
    "BICI_HU0001_ProductoDigitalBicibles": "NO_INICIADO",
    "EXEQ_HU0001_ExequiasPositiva": "NO_INICIADO",
    "ACCAI_HU0001_Formulario": "NO_INICIADO",
    "IMPU_HU0001_GenerarCertificados": "NO_INICIADO",
    "DOCU_HU0001_BibliotecaListasInteligentes": "NO_INICIADO",
    "BANN_HU0001_AdminBanners": "NO_INICIADO",
    "GEWE_HU0001_AdminEstadoActualizacion": "NO_INICIADO",
}

CATEGORY_RE = re.compile(r"^([A-Za-zÁÉÍÓÚáéíóúñÑ]+)_HU\d+_")

COMBINING_DIACRITICS = re.compile("[" + chr(0x0300) + "-" + chr(0x036F) + "]")


def slugify(s):
    s = s.lower()
    s = unicodedata.normalize("NFD", s)
    s = COMBINING_DIACRITICS.sub("", s)
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def norm(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def to_number(v):
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str) and v.strip() != "":
        try:
            return float(v)
        except ValueError:
            return None
    return None


def to_date(v):
    import datetime

    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def normalizar_fase(raw):
    limpio = raw.replace("▶", "").replace("►", "").strip().upper()
    if limpio.startswith("REQUERIM"):
        return (1, "Requerimientos")
    if limpio.startswith("DISE"):
        return (2, "Diseño")
    if limpio.startswith("DESARROL"):
        return (3, "Desarrollo")
    if limpio.startswith("QA"):
        return (4, "QA")
    if limpio.startswith("PRODUC"):
        return (5, "Producción")
    return None


def category_from_code(code):
    if not code:
        return None
    m = CATEGORY_RE.match(code)
    return m.group(1).upper() if m else None


def leer_requerimientos(wb):
    ws = wb[DASHBOARD_SHEET]
    requerimientos = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        estado_cell, item_cell, nombre_cell, mes_cell, complejidad_cell = (
            row[0].value,
            row[1].value,
            row[2].value,
            row[3].value,
            row[4].value,
        )
        horas_est_cell, horas_ejec_cell, fecha_cobro_cell, notas_cell, hoja_detalle_cell = (
            row[11].value,
            row[12].value,
            row[13].value,
            row[14].value,
            row[15].value,
        )

        item = norm(item_cell)
        nombre = norm(nombre_cell)
        if not item and not nombre:
            continue

        hoja_detalle = norm(hoja_detalle_cell)
        tiene_detalle = hoja_detalle is not None

        estado_explicito = norm(estado_cell)
        if estado_explicito:
            estado = ESTADO_ES_A_ENUM.get(estado_explicito.lower(), "NO_INICIADO")
        elif item and item in ESTADO_HEURISTICO:
            estado = ESTADO_HEURISTICO[item]
        else:
            estado = "NO_INICIADO"

        code = item or nombre
        requerimientos.append(
            {
                "code": code,
                "slug": slugify(item) if item else slugify(nombre or f"req-{row[0].row}"),
                "title": nombre or item or "Sin nombre",
                "category": category_from_code(code),
                "complexity": norm(complejidad_cell),
                "month_label": norm(mes_cell),
                "status": estado,
                "has_detail_tracking": tiene_detalle,
                "estimated_hours": to_number(horas_est_cell) or 0,
                "executed_hours": to_number(horas_ejec_cell) or 0,
                "billing_date": norm(fecha_cobro_cell),
                "notes": norm(notas_cell),
                "hoja_detalle": hoja_detalle,
            }
        )
    return requerimientos


def leer_tareas_hoja(wb, hoja_nombre):
    if hoja_nombre not in wb.sheetnames:
        print(f"  ADVERTENCIA: hoja de detalle '{hoja_nombre}' no encontrada en el workbook")
        return []
    ws = wb[hoja_nombre]
    tareas = []
    fase_actual = None
    sort_order = 0

    for row in ws.iter_rows(min_row=4, values_only=False):
        col_a = norm(row[0].value)
        if not col_a:
            continue
        if col_a.upper() == "TOTALES":
            continue
        if "▶" in col_a or "►" in col_a:
            fase = normalizar_fase(col_a)
            fase_actual = fase
            continue
        if fase_actual is None:
            continue

        tarea_nombre = norm(row[2].value)
        if not tarea_nombre:
            continue

        tareas.append(
            {
                "phase_number": fase_actual[0],
                "phase_name": fase_actual[1],
                "task_name": tarea_nombre,
                "detail": norm(row[3].value),
                "status": norm(row[4].value) or "Pendiente",
                "estimated_hours": to_number(row[5].value),
                "due_date": to_date(row[6].value),
                "completed_date": to_date(row[7].value),
                "milestone": norm(row[8].value),
                "notes": norm(row[9].value),
                "blockers": norm(row[10].value),
                "sort_order": sort_order,
            }
        )
        sort_order += 1

    return tareas


def jsonable(req_or_tarea):
    """Convierte date -> isoformat (JSON no soporta datetime.date directamente)."""
    out = {}
    for k, v in req_or_tarea.items():
        if hasattr(v, "isoformat"):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return out


def migrar(supabase, reset=False):
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    requerimientos = leer_requerimientos(wb)

    proyecto = supabase.table("projects").select("id").eq("slug", PROJECT_SLUG).execute()
    if not proyecto.data:
        raise RuntimeError(
            f"No existe el proyecto '{PROJECT_SLUG}' -- corre supabase/schema.sql primero (incluye la seed)."
        )
    project_id = proyecto.data[0]["id"]

    if reset:
        print(f"--reset: borrando requirements existentes del proyecto {PROJECT_SLUG}...")
        supabase.table("requirements").delete().eq("project_id", project_id).execute()

    estados_tareas_vistos = set()
    conteo_tareas_por_requerimiento = {}

    for req in requerimientos:
        hoja_detalle = req.pop("hoja_detalle")
        payload = jsonable({**req, "project_id": project_id})
        resp = (
            supabase.table("requirements")
            .upsert(payload, on_conflict="project_id,code")
            .execute()
        )
        requirement_id = resp.data[0]["id"]

        if not req["has_detail_tracking"]:
            continue

        tareas = leer_tareas_hoja(wb, hoja_detalle)
        conteo_tareas_por_requerimiento[req["code"]] = len(tareas)
        fechas_due = [t["due_date"] for t in tareas if t["due_date"]]

        for t in tareas:
            estados_tareas_vistos.add(t["status"])
            payload_tarea = jsonable({**t, "requirement_id": requirement_id})
            supabase.table("requirement_tasks").upsert(
                payload_tarea, on_conflict="requirement_id,phase_number,task_name"
            ).execute()

        deadline = max(fechas_due) if fechas_due else None
        supabase.table("requirements").update(
            {"deadline": deadline.isoformat() if deadline else None}
        ).eq("id", requirement_id).execute()

    return {
        "total_requerimientos": len(requerimientos),
        "con_detalle": sum(1 for r in requerimientos if r["has_detail_tracking"]),
        "estados_tareas_vistos": sorted(estados_tareas_vistos),
        "conteo_tareas_por_requerimiento": conteo_tareas_por_requerimiento,
    }


def reporte_verificacion(supabase, resultado):
    print("\n" + "=" * 70)
    print("REPORTE DE VERIFICACION")
    print("=" * 70)

    total_db = len(
        supabase.table("requirements").select("id", count="exact").execute().data
    )
    con_detalle_resp = (
        supabase.table("requirements")
        .select("id")
        .eq("has_detail_tracking", True)
        .execute()
    )
    con_detalle_db = len(con_detalle_resp.data)

    print(f"1. count(requirements) en BD = {total_db}  (esperado 28)")
    if total_db != 28:
        print("   *** FALLA: el conteo no coincide con 28 -- revisar antes de continuar. ***")

    print(f"2. count(requirements con detalle) en BD = {con_detalle_db}  (esperado 7)")
    if con_detalle_db != 7:
        print("   *** FALLA: el conteo de 'con detalle' no coincide con 7. ***")

    print("3. Conteo de tareas por requerimiento con detalle (comparar a ojo contra el Excel):")
    for code, n in resultado["conteo_tareas_por_requerimiento"].items():
        print(f"   - {code}: {n} tareas")

    print(f"4. Set de valores distintos de 'status' en tareas: {resultado['estados_tareas_vistos']}")

    print("5. Spot-check -- corre esto manualmente en el SQL Editor y compara contra el Excel abierto:")
    print("   select code, status, estimated_hours, executed_hours, deadline from requirements limit 5;")

    if total_db == 28 and con_detalle_db == 7:
        print("\nMigracion verificada OK.")
    else:
        print("\nMigracion con inconsistencias -- NO continuar a la Fase A sin resolver esto.")
        sys.exit(1)


def main():
    reset = "--reset" in sys.argv

    if not EXCEL_PATH.exists():
        print(f"ERROR: no se encontro el archivo en {EXCEL_PATH}")
        sys.exit(1)

    url = os.environ.get("SUPABASE_URL")
    secret_key = os.environ.get("SUPABASE_SECRET_KEY")
    if not url or not secret_key:
        print("ERROR: faltan las variables de entorno SUPABASE_URL y/o SUPABASE_SECRET_KEY.")
        sys.exit(1)

    supabase = create_client(url, secret_key)
    resultado = migrar(supabase, reset=reset)
    reporte_verificacion(supabase, resultado)


if __name__ == "__main__":
    main()
